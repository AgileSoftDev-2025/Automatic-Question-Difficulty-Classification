from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import FileResponse, JsonResponse, HttpResponse, Http404
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Sum, Count, Avg
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.conf import settings
import os
import json
from datetime import datetime, timedelta
import mimetypes
import logging

# Import ML model classifier
from .ml_model import classify_questions_batch, get_classifier, classify_question

logger = logging.getLogger(__name__)

# Rest of your imports for PDF generation...
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    Spacer, PageBreak, Image, KeepTogether
)
from reportlab.pdfgen import canvas
from io import BytesIO

# Uncomment when models are ready
# from .models import Classification, Question

def redirect_to_main_home(request):
    """
    Redirect /klasifikasi/ to main home page
    Since home functionality is in the 'soal' app
    """
    return redirect('soal:home')
@login_required
@require_http_methods(["GET"])
def hasil_klasifikasi(request, pk=None):
    """
    Display classification results page with ML model predictions
    """
    try:
        # WHEN MODEL IS READY - Replace with:
        """
        classification = get_object_or_404(
            Classification,
            pk=pk,
            user=request.user
        )
        
        questions = classification.questions.select_related('classification').order_by('question_number')
        questions_text = [q.question_text for q in questions]
        """
        
        # DUMMY DATA - Replace this with actual extracted questions
        questions_text = [
            'Apa yang dimaksud dengan algoritma dalam pemrograman?',
            'Jelaskan perbedaan antara bahasa pemrograman tingkat tinggi dan tingkat rendah',
            'Gunakan struktur percabangan untuk membuat program sederhana menentukan bilangan genap atau ganjil',
            'Analisis penyebab program tidak berjalan dengan benar meskipun sintaks sudah benar',
            'Evaluasi efektivitas penggunaan bubble sort dibandingkan insertion sort untuk dataset besar',
            'Rancang algoritma untuk menghitung total belanja dengan diskon dan pajak menggunakan bahasa pemrograman pilihanmu',
            'Sebutkan tiga jenis topologi jaringan komputer',
            'Jelaskan fungsi dari IP address dalam jaringan komputer',
            'Implementasikan konfigurasi jaringan sederhana menggunakan aplikasi Packet Tracer',
            'Buat rancangan sistem absensi mahasiswa berbasis web dengan mempertimbangkan keamanan data pengguna',
            'Apa itu variabel dalam pemrograman?',
            'Jelaskan konsep loop dalam pemrograman',
            'Terapkan konsep array untuk menyimpan data mahasiswa',
            'Bandingkan efisiensi algoritma sorting berbeda',
            'Evaluasi keamanan sistem login yang ada',
            'Desain database untuk sistem perpustakaan',
            'Sebutkan jenis-jenis operator dalam pemrograman',
            'Jelaskan perbedaan GET dan POST method',
            'Implementasikan CRUD operations untuk data siswa',
            'Analisis performa aplikasi web Anda',
            'Rancang arsitektur microservices untuk e-commerce',
        ]
        
        try:
            # Classify all questions using ML model
            logger.info(f"Starting ML classification for {len(questions_text)} questions")
            
            predictions = classify_questions_batch(
                questions_text,
                translate=True,  # Translate Indonesian to English
                batch_size=8
            )
            
            # Build questions data with ML predictions
            questions_data = []
            for i, (text, pred) in enumerate(zip(questions_text, predictions), 1):
                questions_data.append({
                    'question': text,
                    'level': pred['category'],  # C1-C6
                    'index': i,
                    'confidence': pred['confidence'] * 100  # Convert to percentage
                })
            
            logger.info(f"Successfully classified {len(questions_data)} questions using ML model")
            
        except Exception as e:
            logger.error(f"ML Classification error: {str(e)}", exc_info=True)
            messages.warning(
                request, 
                'Automatic classification encountered an issue. Displaying questions without classification.'
            )
            
            # Fallback: show questions without classification
            questions_data = [
                {
                    'question': text,
                    'level': 'C2',  # Default
                    'index': i,
                    'confidence': 0
                }
                for i, text in enumerate(questions_text, 1)
            ]
        
        filename = 'SOAL LATIHAN UTS.pdf'
        total_questions = len(questions_data)
        
        context = {
            'filename': filename,
            'file_url': '#',
            'questions': questions_data,
            'labels': ['C1', 'C2', 'C3', 'C4', 'C5', 'C6'],
            'total_questions': total_questions,
            'classification_id': pk or 1,
            'classification_date': timezone.now().strftime('%d/%m/%Y'),
        }
        
        return render(request, 'klasifikasi/hasilKlasifikasi.html', context)
        
    except Exception as e:
        logger.error(f"Error in hasil_klasifikasi view: {str(e)}", exc_info=True)
        messages.error(request, 'An error occurred while loading classification results.')
        return redirect('klasifikasi:history')


@login_required
@csrf_protect
@require_POST
def update_question_classification(request, pk):
    """
    AJAX endpoint to update individual question classification
    """
    try:
        data = json.loads(request.body)
        question_id = data.get('question_id')
        new_category = data.get('category')
        
        if not question_id or not new_category:
            return JsonResponse({
                'success': False,
                'error': 'Missing required fields'
            }, status=400)
        
        valid_categories = ['C1', 'C2', 'C3', 'C4', 'C5', 'C6']
        if new_category not in valid_categories:
            return JsonResponse({
                'success': False,
                'error': f'Invalid category. Must be one of: {", ".join(valid_categories)}'
            }, status=400)
        
        # WHEN MODEL IS READY:
        """
        classification = get_object_or_404(Classification, pk=pk, user=request.user)
        question = get_object_or_404(Question, id=question_id, classification=classification)
        
        old_category = question.category
        question.category = new_category
        question.is_manually_classified = True
        question.save()
        
        classification.recalculate_counts()
        """
        
        logger.info(
            f"Question {question_id} updated to {new_category} "
            f"by user {request.user.username} in classification {pk}"
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Classification updated successfully',
            'question_id': question_id,
            'new_category': new_category
        })
        
    except Exception as e:
        logger.error(f"Error updating question classification: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An internal error occurred'
        }, status=500)


@login_required
@require_http_methods(["GET", "POST"])
def home(request):
    """Display home page with file upload"""
    recent_history = []
    
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        
        if not uploaded_file:
            messages.error(request, 'Please select a file to upload.')
            return redirect('klasifikasi:home')
        
        # File validation
        allowed_extensions = ['.pdf', '.doc', '.docx']
        file_extension = os.path.splitext(uploaded_file.name)[1].lower()
        
        if file_extension not in allowed_extensions:
            messages.error(
                request, 
                f'File format not supported. Only {", ".join(allowed_extensions)} files are allowed.'
            )
            return redirect('klasifikasi:home')
        
        # Validate file size (max 10MB)
        max_size = 10 * 1024 * 1024
        if uploaded_file.size > max_size:
            size_mb = uploaded_file.size / (1024 * 1024)
            messages.error(
                request, 
                f'File size too large. Maximum 10MB. Your file is {size_mb:.1f}MB.'
            )
            return redirect('klasifikasi:home')
        
        try:
            # TODO: Extract questions from file
            # TODO: Run ML classification
            # TODO: Save to database
            
            messages.success(
                request, 
                f'File "{uploaded_file.name}" uploaded successfully. Processing...'
            )
            return redirect('klasifikasi:history')
            
        except Exception as e:
            logger.error(f"Error processing file upload: {str(e)}", exc_info=True)
            messages.error(request, 'An error occurred while processing your file.')
            return redirect('klasifikasi:home')
    
    # GET request
    try:
        stats = {
            'total': 0,
            'total_questions': 0,
            'avg_processing_time': 0
        }
    except Exception as e:
        logger.error(f"Error fetching statistics: {str(e)}", exc_info=True)
        stats = {}
    
    context = {
        'recent_history': recent_history,
        'stats': stats,
        'is_authenticated': request.user.is_authenticated,
        'max_file_size_mb': 10,
        'allowed_formats': ['PDF', 'DOC', 'DOCX']
    }
    
    return render(request, 'klasifikasi/home.html', context)


@login_required
@require_http_methods(["GET"])
def history_view(request):
    """Display classification history - USING HARDCODED DATA"""
    try:
        search_query = request.GET.get('search', '').strip()
        sort_by = request.GET.get('sort', 'date-desc')
        
        # HARDCODED DUMMY DATA
        classifications_dummy = [
            {
                'id': 1,
                'filename': 'SOAL LATIHAN UTS.pdf',
                'total_questions': 21,
                'created_at': '09/11/2025',
                'q1_count': 4,
                'q2_count': 5,
                'q3_count': 4,
                'q4_count': 2,
                'q5_count': 2,
                'q6_count': 4,
                'status': 'completed',
            },
            {
                'id': 2,
                'filename': 'Soal UTS Semester Ganjil.pdf',
                'total_questions': 10,
                'created_at': '15/10/2025',
                'q1_count': 2,
                'q2_count': 2,
                'q3_count': 1,
                'q4_count': 3,
                'q5_count': 2,
                'q6_count': 0,
                'status': 'completed',
            },
        ]
        
        filtered_classifications = classifications_dummy.copy()
        
        # Apply search filter
        if search_query:
            filtered_classifications = [
                c for c in filtered_classifications 
                if search_query.lower() in c['filename'].lower()
            ]
        
        # Sort
        if sort_by == 'date-desc':
            filtered_classifications.sort(
                key=lambda x: datetime.strptime(x['created_at'], '%d/%m/%Y'),
                reverse=True
            )
        elif sort_by == 'date-asc':
            filtered_classifications.sort(
                key=lambda x: datetime.strptime(x['created_at'], '%d/%m/%Y')
            )
        
        # Statistics
        total_classifications = len(filtered_classifications)
        total_questions = sum(c['total_questions'] for c in filtered_classifications)
        last_activity = filtered_classifications[0]['created_at'] if filtered_classifications else 'N/A'
        
        # Pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(filtered_classifications, 10)
        
        try:
            page_obj = paginator.page(page)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        
        context = {
            'classifications': page_obj,
            'total_classifications': total_classifications,
            'total_questions': total_questions,
            'last_activity': last_activity,
            'search_query': search_query,
            'sort_by': sort_by,
        }
        
        return render(request, 'klasifikasi/history.html', context)
        
    except Exception as e:
        logger.error(f"Error in history_view: {str(e)}", exc_info=True)
        return redirect('klasifikasi:home')



@login_required
@csrf_protect
@require_POST
def delete_classification(request, pk):
    """
    Delete classification with proper error handling and security
    """
    try:
        # WHEN MODEL IS READY:
        """
        classification = get_object_or_404(
            Classification, 
            pk=pk, 
            user=request.user
        )
        
        filename = classification.filename
        
        # Delete the classification (files will be deleted by signal)
        classification.delete()
        
        logger.info(f"Classification {pk} ({filename}) deleted by user {request.user.username}")
        
        # Return JSON response for AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True, 
                'message': f'"{filename}" deleted successfully'
            })
        
        messages.success(request, f'Classification "{filename}" deleted successfully.')
        """
        
        # DUMMY response
        logger.info(f"Classification {pk} would be deleted by user {request.user.username}")
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True, 
                'message': 'Classification deleted successfully'
            })
        
        messages.success(request, 'Classification deleted successfully.')
        return redirect('klasifikasi:history')
        
    except Http404:
        error_msg = 'Classification not found or you do not have permission to delete it.'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=404)
        messages.error(request, error_msg)
        return redirect('klasifikasi:history')
        
    except Exception as e:
        logger.error(f"Error deleting classification {pk}: {str(e)}", exc_info=True)
        error_msg = 'An error occurred while deleting the classification.'
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': error_msg}, status=500)
        
        messages.error(request, error_msg)
        return redirect('klasifikasi:history')


def add_page_number(canvas_obj, doc):
    """Add page numbers to PDF"""
    page_num = canvas_obj.getPageNumber()
    text = f"Page {page_num}"
    canvas_obj.saveState()
    canvas_obj.setFont('Helvetica', 9)
    canvas_obj.setFillColor(colors.grey)
    canvas_obj.drawRightString(
        letter[0] - 0.75 * inch,
        0.5 * inch,
        text
    )
    canvas_obj.restoreState()


@login_required
@require_http_methods(["GET"])
def download_report(request, pk):
    """
    Download classification report as PDF
    Currently uses dummy data - will be replaced with model data
    """
    try:
        # DUMMY DATA - Replace with actual model data when ready
        """
        # WHEN MODEL IS READY - Uncomment this:
        classification = get_object_or_404(
            Classification, 
            pk=pk, 
            user=request.user
        )
        
        if not classification.result_file:
            messages.error(request, 'Report file not available.')
            return redirect('klasifikasi:history')
        
        file_path = classification.result_file.path
        
        if not os.path.exists(file_path):
            messages.error(request, 'Report file not found on server.')
            return redirect('klasifikasi:history')
        """
        
        # DUMMY DATA for testing
        classification_data = {
            'id': pk,
            'filename': 'SOAL LATIHAN UTS.pdf',
            'total_questions': 21,
            'created_at': timezone.now(),
            'q1_count': 4,
            'q2_count': 5,
            'q3_count': 4,
            'q4_count': 2,
            'q5_count': 2,
            'q6_count': 4,
            'status': 'completed',
            'user': request.user,
        }
        
        questions_data = [
            {
                'question_number': 1,
                'question_text': 'Apa yang dimaksud dengan algoritma dalam pemrograman?',
                'category': 'C1',
                'confidence_score': 0.95,
            },
            {
                'question_number': 2,
                'question_text': 'Jelaskan perbedaan antara bahasa pemrograman tingkat tinggi dan tingkat rendah',
                'category': 'C2',
                'confidence_score': 0.89,
            },
            {
                'question_number': 3,
                'question_text': 'Gunakan struktur percabangan untuk membuat program sederhana menentukan bilangan genap atau ganjil',
                'category': 'C3',
                'confidence_score': 0.92,
            },
            {
                'question_number': 4,
                'question_text': 'Analisis penyebab program tidak berjalan dengan benar meskipun sintaks sudah benar',
                'category': 'C4',
                'confidence_score': 0.87,
            },
            {
                'question_number': 5,
                'question_text': 'Evaluasi efektivitas penggunaan bubble sort dibandingkan insertion sort untuk dataset besar',
                'category': 'C5',
                'confidence_score': 0.91,
            },
            {
                'question_number': 6,
                'question_text': 'Rancang algoritma untuk menghitung total belanja dengan diskon dan pajak menggunakan bahasa pemrograman pilihanmu',
                'category': 'C6',
                'confidence_score': 0.88,
            },
            {
                'question_number': 7,
                'question_text': 'Sebutkan tiga jenis topologi jaringan komputer',
                'category': 'C1',
                'confidence_score': 0.94,
            },
            {
                'question_number': 8,
                'question_text': 'Jelaskan fungsi dari IP address dalam jaringan komputer',
                'category': 'C2',
                'confidence_score': 0.90,
            },
            {
                'question_number': 9,
                'question_text': 'Implementasikan konfigurasi jaringan sederhana menggunakan aplikasi Packet Tracer',
                'category': 'C3',
                'confidence_score': 0.86,
            },
            {
                'question_number': 10,
                'question_text': 'Buat rancangan sistem absensi mahasiswa berbasis web dengan mempertimbangkan keamanan data pengguna',
                'category': 'C6',
                'confidence_score': 0.93,
            },
            {
                'question_number': 11,
                'question_text': 'Apa itu variabel dalam pemrograman?',
                'category': 'C1',
                'confidence_score': 0.96,
            },
            {
                'question_number': 12,
                'question_text': 'Jelaskan konsep loop dalam pemrograman',
                'category': 'C2',
                'confidence_score': 0.88,
            },
            {
                'question_number': 13,
                'question_text': 'Terapkan konsep array untuk menyimpan data mahasiswa',
                'category': 'C3',
                'confidence_score': 0.85,
            },
            {
                'question_number': 14,
                'question_text': 'Bandingkan efisiensi algoritma sorting berbeda',
                'category': 'C4',
                'confidence_score': 0.90,
            },
            {
                'question_number': 15,
                'question_text': 'Evaluasi keamanan sistem login yang ada',
                'category': 'C5',
                'confidence_score': 0.87,
            },
            {
                'question_number': 16,
                'question_text': 'Desain database untuk sistem perpustakaan',
                'category': 'C6',
                'confidence_score': 0.89,
            },
            {
                'question_number': 17,
                'question_text': 'Sebutkan jenis-jenis operator dalam pemrograman',
                'category': 'C1',
                'confidence_score': 0.95,
            },
            {
                'question_number': 18,
                'question_text': 'Jelaskan perbedaan GET dan POST method',
                'category': 'C2',
                'confidence_score': 0.91,
            },
            {
                'question_number': 19,
                'question_text': 'Implementasikan CRUD operations untuk data siswa',
                'category': 'C3',
                'confidence_score': 0.88,
            },
            {
                'question_number': 20,
                'question_text': 'Analisis performa aplikasi web Anda',
                'category': 'C4',
                'confidence_score': 0.86,
            },
            {
                'question_number': 21,
                'question_text': 'Rancang arsitektur microservices untuk e-commerce',
                'category': 'C6',
                'confidence_score': 0.92,
            },
        ]
        
        # Generate PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch,
        )
        
        # Container for PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        )
        
        subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#374151'),
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#374151'),
            spaceAfter=6,
        )
        
        # Title
        title = Paragraph("BLOOMERS Classification Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Report Information
        info_data = [
            ['File Name:', classification_data['filename']],
            ['Classification ID:', f"#{classification_data['id']}"],
            ['Generated On:', classification_data['created_at'].strftime('%d/%m/%Y %H:%M')],
            ['Total Questions:', str(classification_data['total_questions'])],
            ['User:', request.user.username],
        ]
        
        info_table = Table(info_data, colWidths=[1.5*inch, 4.5*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#374151')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#6b7280')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Distribution Summary
        elements.append(Paragraph("Classification Distribution", heading_style))
        
        # Calculate percentages
        total = classification_data['total_questions']
        distribution_data = [
            ['Category', 'Level', 'Count', 'Percentage'],
            ['C1', 'Remember', str(classification_data['q1_count']), 
             f"{(classification_data['q1_count']/total*100):.1f}%"],
            ['C2', 'Understand', str(classification_data['q2_count']), 
             f"{(classification_data['q2_count']/total*100):.1f}%"],
            ['C3', 'Apply', str(classification_data['q3_count']), 
             f"{(classification_data['q3_count']/total*100):.1f}%"],
            ['C4', 'Analyze', str(classification_data['q4_count']), 
             f"{(classification_data['q4_count']/total*100):.1f}%"],
            ['C5', 'Evaluate', str(classification_data['q5_count']), 
             f"{(classification_data['q5_count']/total*100):.1f}%"],
            ['C6', 'Create', str(classification_data['q6_count']), 
             f"{(classification_data['q6_count']/total*100):.1f}%"],
        ]
        
        dist_table = Table(distribution_data, colWidths=[1*inch, 1.5*inch, 1*inch, 1.5*inch])
        dist_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(dist_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Questions Detail
        elements.append(PageBreak())
        elements.append(Paragraph("Detailed Question Classification", heading_style))
        elements.append(Spacer(1, 0.1*inch))
        
        # Category colors
        category_colors = {
            'C1': colors.HexColor('#dcfce7'),
            'C2': colors.HexColor('#dbeafe'),
            'C3': colors.HexColor('#fef3c7'),
            'C4': colors.HexColor('#fed7aa'),
            'C5': colors.HexColor('#fecaca'),
            'C6': colors.HexColor('#e9d5ff'),
        }
        
        for question in questions_data:
            # Question header
            q_header = Paragraph(
                f"<b>Question {question['question_number']}</b> - "
                f"Category: {question['category']} | "
                f"Confidence: {question['confidence_score']*100:.1f}%",
                subheading_style
            )
            
            # Question text
            q_text = Paragraph(question['question_text'], normal_style)
            
            # Create a box for each question
            q_data = [
                [q_header],
                [q_text],
            ]
            
            q_table = Table(q_data, colWidths=[6*inch])
            q_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), category_colors.get(question['category'], colors.lightgrey)),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('BOX', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('LEFTPADDING', (0, 0), (-1, -1), 12),
                ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ]))
            
            elements.append(KeepTogether([q_table, Spacer(1, 0.15*inch)]))
        
        # Footer information
        elements.append(PageBreak())
        elements.append(Spacer(1, 0.5*inch))
        
        footer_text = Paragraph(
            "<b>About Bloom's Taxonomy Levels:</b><br/><br/>"
            "<b>C1 (Remember):</b> Recall facts and basic concepts<br/>"
            "<b>C2 (Understand):</b> Explain ideas or concepts<br/>"
            "<b>C3 (Apply):</b> Use information in new situations<br/>"
            "<b>C4 (Analyze):</b> Draw connections among ideas<br/>"
            "<b>C5 (Evaluate):</b> Justify a decision or course of action<br/>"
            "<b>C6 (Create):</b> Produce new or original work<br/><br/>"
            "<i>This report was automatically generated by BLOOMERS Classification System.</i>",
            normal_style
        )
        elements.append(footer_text)
        
        # Build PDF
        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
        
        # Get PDF from buffer
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create response
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"classification_report_{pk}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(
            f"Report downloaded: Classification {pk} "
            f"by {request.user.username}"
        )
        
        return response
        
    except ImportError as e:
        logger.error(f"ReportLab not installed: {str(e)}")
        messages.error(
            request, 
            'PDF generation library not available. Please contact administrator.'
        )
        return redirect('klasifikasi:history')
        
    except Exception as e:
        logger.error(f"Error generating report {pk}: {str(e)}", exc_info=True)
        messages.error(request, 'Error generating report. Please try again.')
        return redirect('klasifikasi:history')


@login_required
@require_http_methods(["GET"])
def get_classification_stats(request):
    """
    API endpoint for classification statistics (AJAX)
    Returns comprehensive statistics about user's classifications
    """
    try:
        # DUMMY data
        stats = {
            'total_classifications': 3,
            'total_questions': 46,
            'completed': 3,
            'processing': 0,
            'failed': 0,
            'categories': {
                'C1': 9,
                'C2': 11,
                'C3': 7,
                'C4': 9,
                'C5': 6,
                'C6': 4,
            },
            'recent_activity': [
                {
                    'id': 1,
                    'filename': 'SOAL LATIHAN UTS.pdf',
                    'date': '09/11/2025',
                    'questions': 21,
                    'status': 'completed'
                },
                {
                    'id': 2,
                    'filename': 'Soal UTS Semester Ganjil.pdf',
                    'date': '15/10/2025',
                    'questions': 10,
                    'status': 'completed'
                },
                {
                    'id': 3,
                    'filename': 'Quiz Matematika.pdf',
                    'date': '22/09/2025',
                    'questions': 15,
                    'status': 'completed'
                }
            ]
        }
        
        # WHEN MODEL IS READY:
        """
        from django.db.models import Sum, Count, Q
        
        classifications = Classification.objects.filter(user=request.user)
        
        # Aggregate statistics
        aggregates = classifications.aggregate(
            total=Count('id'),
            total_questions=Sum('total_questions'),
            completed=Count('id', filter=Q(status='completed')),
            processing=Count('id', filter=Q(status='processing')),
            failed=Count('id', filter=Q(status='failed')),
            total_c1=Sum('q1_count'),
            total_c2=Sum('q2_count'),
            total_c3=Sum('q3_count'),
            total_c4=Sum('q4_count'),
            total_c5=Sum('q5_count'),
            total_c6=Sum('q6_count'),
        )
        
        stats = {
            'total_classifications': aggregates['total'] or 0,
            'total_questions': aggregates['total_questions'] or 0,
            'completed': aggregates['completed'] or 0,
            'processing': aggregates['processing'] or 0,
            'failed': aggregates['failed'] or 0,
            'categories': {
                'C1': aggregates['total_c1'] or 0,
                'C2': aggregates['total_c2'] or 0,
                'C3': aggregates['total_c3'] or 0,
                'C4': aggregates['total_c4'] or 0,
                'C5': aggregates['total_c5'] or 0,
                'C6': aggregates['total_c6'] or 0,
            },
            'recent_activity': [
                {
                    'id': c.id,
                    'filename': c.filename,
                    'date': c.formatted_created_at,
                    'questions': c.total_questions,
                    'status': c.status
                }
                for c in classifications.order_by('-created_at')[:5]
            ]
        }
        """
        
        return JsonResponse(stats)
        
    except Exception as e:
        logger.error(f"Error fetching classification stats: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to fetch statistics'
        }, status=500)


@login_required
@csrf_protect
@require_POST
def bulk_delete_classifications(request):
    """
    Delete multiple classifications at once
    Enhanced with transaction support and better error handling
    """
    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])
        
        if not ids:
            return JsonResponse({
                'success': False,
                'error': 'No classification IDs provided'
            }, status=400)
        
        if not isinstance(ids, list):
            return JsonResponse({
                'success': False,
                'error': 'IDs must be provided as a list'
            }, status=400)
        
        # Validate all IDs are integers
        try:
            ids = [int(id_val) for id_val in ids]
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'error': 'All IDs must be valid integers'
            }, status=400)
        
        # WHEN MODEL IS READY:
        """
        from django.db import transaction
        
        with transaction.atomic():
            classifications = Classification.objects.filter(
                pk__in=ids,
                user=request.user
            ).select_for_update()
            
            deleted_count = 0
            filenames = []
            
            for classification in classifications:
                filenames.append(classification.filename)
                classification.delete()  # Files deleted by signal
                deleted_count += 1
            
            logger.info(
                f"Bulk deleted {deleted_count} classifications by user {request.user.username}: "
                f"{', '.join(filenames[:5])}{'...' if len(filenames) > 5 else ''}"
            )
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'{deleted_count} classification(s) deleted successfully'
        })
        """
        
        # DUMMY response
        logger.info(
            f"Would bulk delete {len(ids)} classifications by user {request.user.username}"
        )
        
        return JsonResponse({
            'success': True,
            'deleted_count': len(ids),
            'message': f'{len(ids)} classification(s) deleted successfully'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.error(f"Error in bulk delete: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while deleting classifications'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def export_classification_excel(request, pk):
    """
    Export classification results to Excel format
    """
    try:
        # WHEN MODEL IS READY:
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from io import BytesIO
        
        classification = get_object_or_404(
            Classification,
            pk=pk,
            user=request.user
        )
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Classification Results"
        
        # Define styles
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers
        headers = ['No', 'Question', 'Category', 'Level Name', 'Confidence', 'Manually Classified']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add data
        questions = classification.questions.all().order_by('question_number')
        for row_num, question in enumerate(questions, 2):
            ws.cell(row=row_num, column=1, value=question.question_number).border = border
            ws.cell(row=row_num, column=2, value=question.question_text).border = border
            ws.cell(row=row_num, column=3, value=question.category).border = border
            ws.cell(row=row_num, column=4, value=question.category_name).border = border
            ws.cell(row=row_num, column=5, value=question.formatted_confidence).border = border
            ws.cell(row=row_num, column=6, value='Yes' if question.is_manually_classified else 'No').border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        
        # Save to response
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"classification_{classification.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Excel export: {classification.filename} by {request.user.username}")
        
        return response
        """
        
        messages.info(request, 'Excel export functionality will be available soon.')
        return redirect('klasifikasi:hasil_klasifikasi', pk=pk)
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}", exc_info=True)
        messages.error(request, 'Error exporting to Excel.')
        return redirect('klasifikasi:hasil_klasifikasi', pk=pk)


@login_required
@require_http_methods(["GET"])
def get_question_details(request, classification_id, question_id):
    """
    API endpoint to get detailed information about a specific question
    """
    try:
        # WHEN MODEL IS READY:
        """
        classification = get_object_or_404(
            Classification,
            pk=classification_id,
            user=request.user
        )
        
        question = get_object_or_404(
            Question,
            pk=question_id,
            classification=classification
        )
        
        data = {
            'id': question.id,
            'number': question.question_number,
            'text': question.question_text,
            'category': question.category,
            'category_name': question.category_name,
            'category_description': question.category_description,
            'confidence': question.confidence_score,
            'formatted_confidence': question.formatted_confidence,
            'is_manually_classified': question.is_manually_classified,
            'previous_category': question.previous_category,
            'has_choices': question.has_choices,
            'choices': question.choices_list if question.has_choices else [],
            'correct_answer': question.correct_answer,
            'created_at': question.created_at.isoformat(),
            'updated_at': question.updated_at.isoformat()
        }
        
        return JsonResponse(data)
        """
        
        # DUMMY response
        return JsonResponse({
            'id': question_id,
            'number': 1,
            'text': 'Sample question text',
            'category': 'C1',
            'category_name': 'Remember',
            'confidence': 0.95,
            'is_manually_classified': False
        })
        
    except Exception as e:
        logger.error(f"Error fetching question details: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to fetch question details'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def classification_analytics(request, pk):
    """
    Display detailed analytics for a classification
    """
    try:
        # WHEN MODEL IS READY:
        """
        classification = get_object_or_404(
            Classification,
            pk=pk,
            user=request.user
        )
        
        # Get distribution data
        distribution = classification.distribution_percentages
        
        # Get question confidence statistics
        questions = classification.questions.all()
        avg_confidence = questions.aggregate(Avg('confidence_score'))['confidence_score__avg'] or 0
        
        # Get manually classified count
        manual_count = questions.filter(is_manually_classified=True).count()
        
        context = {
            'classification': classification,
            'distribution': distribution,
            'avg_confidence': avg_confidence * 100,
            'manual_count': manual_count,
            'auto_count': classification.total_questions - manual_count
        }
        
        return render(request, 'klasifikasi/analytics.html', context)
        """
        
        messages.info(request, 'Analytics page will be available soon.')
        return redirect('klasifikasi:hasil_klasifikasi', pk=pk)
        
    except Exception as e:
        logger.error(f"Error loading analytics: {str(e)}", exc_info=True)
        messages.error(request, 'Error loading analytics.')
        return redirect('klasifikasi:history')